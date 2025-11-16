#!/usr/bin/env python3
"""Extract and display data from Helistat Excel files"""

import openpyxl
import sys

def read_excel_file(filepath):
    """Read and display contents of an Excel file"""
    print(f"\n{'='*80}")
    print(f"Reading: {filepath}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    print(f"Sheet names: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}")

        # Read first 50 rows
        max_rows = min(50, ws.max_row)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), 1):
            # Filter out completely empty rows
            if any(cell is not None for cell in row):
                # Format the row, showing only non-None values compactly
                row_data = []
                for col_idx, cell in enumerate(row, 1):
                    if cell is not None:
                        row_data.append(f"[{col_idx}] {cell}")

                if row_data:
                    print(f"Row {row_idx}: {' | '.join(row_data[:10])}")  # Limit to 10 columns for readability

        print("")

if __name__ == "__main__":
    files = [
        "/home/user/z2i/BAR/Helistat_v4.0.xlsx",
        "/home/user/z2i/BAR/Micro Helistat_v1.0.xlsx"
    ]

    for f in files:
        try:
            read_excel_file(f)
        except Exception as e:
            print(f"Error reading {f}: {e}")
